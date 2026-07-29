"""
data_cleaning.py - 麦当劳内容排行榜：数据清洗
"""

import json
import logging
import re
import datetime
import numpy as np
import pandas as pd
from io import BytesIO

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
# 列名常量（参考 mcd-reach-trend/data.py，便于跨源文件兼容）
# ═══════════════════════════════════════════════════════════════
DATE_COL = "发送日期"
COL_ALIASES = {
    DATE_COL: ["日期", "send", "date", "send_date", "send time", "sendtime"],
    "渠道": ["渠道", "channel"],
    "计划类型": ["计划类型", "plan_type", "plan type"],
    "plan_id": ["plan_id", "plan id", "planid"],
    "unit_id": ["unit_id", "unit id", "unitid"],
    "message_id": ["message_id", "message id", "messageid", "msg_id"],
    "plan名称": ["plan名称", "plan_name", "plan name"],
    "owner": ["owner", "预算owner", "预算 owner", "预算_owner", "bu"],
    "是否用券": ["是否用券", "coupon"],
    "预计触达": ["预计触达", "exp_reach", "expected reach"],
    "触达成功": ["触达成功", "reach", "reach_success"],
    "点击人次": ["点击人次", "click"],
    "点击后下单人次": ["点击后下单", "点击后下单人次", "post_click", "下单人次", "order_conv"],
    "订单GC": ["gc", "订单gc", "order_gc"],
    "订单Sales": ["sales", "订单sales", "order_sales"],
    # 消息 JSON 列：上游可能叫「消息内容」或英文别名，2026-07 数据源在 O 列
    # 新增了 Message ID，改靠列名别名而非 iloc[:, 14] 定位
    "消息内容": ["消息内容", "message_content", "msg_content", "content_json"],
}

# Excel 日期序列号识别范围：1~80000 覆盖 1900~2119 年
_EXCEL_SERIAL_MIN = 1
_EXCEL_SERIAL_MAX = 80000

# ID 列：必须按字符串处理
#   Message ID 是 18 位数字（如 592304644657221633），超出 float64 的 2^53 安全整数范围，
#   一旦被当数值转换，不同 ID 会舍入成同一个值，聚合时撞车
#   Unit ID 的 204/1054 ≈ 20% 行是字符串 "[NULL]" 占位 —— 业务上等同空值
#   （没有 Unit 的旧式 Plan），按 Unit 数去重时忽略。判断见
#   scoring._normalize_unit_ids（统一成 NaN 后 dropna 默认行为）
ID_COLS = ("plan_id", "unit_id", "message_id")

# JSON 消息列的标准名（_map_columns 会把别名归一成这个；找不到则读不到 JSON）
MSG_COL = "消息内容"


def _normalize_id_columns(df: pd.DataFrame) -> pd.DataFrame:
    """ID 列统一为字符串并去掉首尾空白（上游 Message ID 带尾部制表符）"""
    for col in ID_COLS:
        if col in df.columns:
            s = df[col].astype(str).str.strip()
            df[col] = s.mask(s.isin(["nan", "None", "NaT", "<NA>"]), "")
    return df


def _locate_message_col(df: pd.DataFrame):
    """定位 JSON 消息列（_map_columns 已把别名归一成标准名），找不到返回 None。

    历史实现写死 iloc[:, 14]（O 列），上游一旦插列就会读错列且不报错。
    现在依赖 COL_ALIASES 里的别名映射：如果上游将来再用新名字，加进
    「消息内容」别名即可，无需再改这里。
    """
    return MSG_COL if MSG_COL in df.columns else None



def _map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """列名模糊匹配：把源列名映射到标准中文列名（提升跨源文件兼容性）"""
    if df.columns.empty:
        return df
    existing = set(df.columns)
    rename = {}
    for target, keys in COL_ALIASES.items():
        if target in existing:
            continue  # 已有标准列名，不重复映射
        for c in df.columns:
            if c in rename:
                continue
            cl = str(c).lower().strip().replace(" ", "").replace("_", "")
            for k in keys:
                kl = str(k).lower().replace(" ", "").replace("_", "")
                if kl in cl:
                    rename[c] = target
                    break
    if rename:
        df = df.rename(columns=rename)
    return df


def _parse_date_column(series: pd.Series) -> pd.Series:
    """智能解析日期列：datetime 对象 / 字符串 / Excel 数字序列号三种情况都能转。
    字符串会做多格式 fallback（ISO / 斜杠 / 点号 / 中文年月日 等）。"""
    if series is None or len(series) == 0:
        return series
    # 已经是 datetime64，直接返回
    if pd.api.types.is_datetime64_any_dtype(series):
        return series
    # 取第一个非空样本判断类型
    non_null = series.dropna()
    if non_null.empty:
        return series
    sample = non_null.iloc[0]
    # 1) datetime 对象（Excel 单元格为日期格式 → openpyxl 读出 Timestamp/datetime）
    if isinstance(sample, (pd.Timestamp, datetime.datetime, datetime.date)):
        return pd.to_datetime(series, errors="coerce")
    # 2) 数字（Excel 日期序列号：1900-01-01=1，1899-12-30 起点可绕过 Excel 1900 闰年 bug）
    if isinstance(sample, (int, float, np.integer, np.floating)) and not isinstance(sample, bool):
        if _EXCEL_SERIAL_MIN <= float(sample) <= _EXCEL_SERIAL_MAX:
            return pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce")
        return series
    # 3) 字符串 - 先试默认 to_datetime，失败/缺失多则 fallback 显式格式
    parsed = pd.to_datetime(series, errors="coerce", format="mixed")
    if parsed.notna().sum() < non_null.shape[0]:
        # 默认解析器未能覆盖全部，尝试常见中文/Excel 日期格式
        for fmt in (
            "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
            "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
            "%d/%m/%Y", "%m/%d/%Y",
        ):
            try:
                fallback = pd.to_datetime(series, errors="coerce", format=fmt)
                # 优先保留首次 fallback 能解析的值
                parsed = parsed.fillna(fallback)
                if parsed.notna().sum() >= non_null.shape[0]:
                    break
            except Exception:
                continue
    return parsed


def extract_title_from_forms(forms):
    """从 forms 列表中提取标题"""
    if not isinstance(forms, list):
        return None
    for item in forms:
        if item.get('code') == 'thing1' and item.get('value'):
            return item['value']
    for item in forms:
        code = item.get('code', '')
        value = item.get('value')
        if code.startswith('thing') and value:
            return value
    for item in forms:
        code = item.get('code', '')
        value = item.get('value')
        if not code.startswith('time') and value:
            return value
    return None


def extract_text_from_forms(forms):
    """从 forms 列表中提取正文"""
    if not isinstance(forms, list):
        return None
    for item in forms:
        code = item.get('code', '')
        value = item.get('value')
        if code in ['thing5', 'short_thing5'] and value:
            return value
    for item in forms:
        code = item.get('code', '')
        value = item.get('value')
        if code.startswith('thing') and code != 'thing1' and value:
            return value
    return None


def parse_message(raw, strip_question_marks=False):
    """将原始 JSON 消息解析为标题和正文"""
    if pd.isna(raw) or not isinstance(raw, str):
        return pd.Series({'标题': '', '内容': ''})

    try:
        data = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return pd.Series({'标题': '', '内容': ''})

    title = data.get('title')
    if not title:
        title = extract_title_from_forms(data.get('forms'))
    if not title:
        attachments = data.get('attachments')
        if isinstance(attachments, list) and len(attachments) > 0:
            title = attachments[0].get('name', '')

    text = data.get('text')
    if not text:
        text = extract_text_from_forms(data.get('forms'))

    if not title and text:
        first_part = re.split(r'[。！？\n]', str(text).strip())[0].strip()
        if len(first_part) > 0:
            title = first_part
        else:
            title = str(text)[:20]

    title = str(title).strip() if title else ''
    text = str(text).strip() if text else ''

    # 清洗换行符（保留 emoji 和其他 Unicode 字符）
    title = title.replace('\r\n', '').replace('\n', '').replace('\r', '')
    text = text.replace('\r\n', '').replace('\n', '').replace('\r', '')

    # CSV 路径：清理 GBK 编码残留的连续 ? 字符（单个问号保留，避免吞掉合法问号）
    if strip_question_marks:
        title = re.sub(r'\?{2,}', '', title)
        text = re.sub(r'\?{2,}', '', text)

    return pd.Series({'标题': title, '内容': text})


def clean_raw_csv(uploaded_file) -> pd.DataFrame:
    """
    1. 尝试多种编码读取原始 CSV
    2. 列名映射 + ID 列定型
    3. 按列名定位 JSON 消息列（「消息内容」）
    4. 解析 JSON，提取标题和内容
    5. 合并回原 DataFrame，删除原始 JSON 列
    """
    bytes_data = uploaded_file.read()

    # 尝试多种编码
    encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(BytesIO(bytes_data), encoding=enc, on_bad_lines='skip')
            break
        except Exception:
            continue

    if df is None:
        raise ValueError("无法读取 CSV 文件，请检查文件格式")

    # 列名模糊匹配（如 send_date → 发送日期），提升跨源兼容性
    df = _map_columns(df)
    df = _normalize_id_columns(df)

    # 按列名定位 JSON 消息列（不再依赖固定位置，上游插列也不会错位）
    msg_col = _locate_message_col(df)
    if msg_col is None:
        raise ValueError("找不到 JSON 消息列，请确认文件中包含「消息内容」列")

    # 执行解析（CSV 路径需要清理 GBK 编码残留的 ?）
    parsed_df = df[msg_col].apply(lambda x: parse_message(x, strip_question_marks=True))

    # 合并回原 DataFrame，删除原始 JSON 列
    df['标题'] = parsed_df['标题']
    df['内容'] = parsed_df['内容']
    df = df.drop(columns=[msg_col])

    # 日期列智能解析（字符串/数字序列号/ datetime 对象 都能识别）
    if DATE_COL in df.columns:
        df[DATE_COL] = _parse_date_column(df[DATE_COL])

    return df


def read_cleaned_csv(uploaded_file) -> pd.DataFrame:
    """读取已清洗的 CSV，自动尝试多种编码"""
    bytes_data = uploaded_file.read()
    for enc in ['utf-8', 'utf-8-sig', 'gbk']:
        try:
            df = pd.read_csv(BytesIO(bytes_data), encoding=enc)
            df = _map_columns(df)
            df = _normalize_id_columns(df)
            if DATE_COL in df.columns:
                df[DATE_COL] = _parse_date_column(df[DATE_COL])
            return df
        except Exception:
            continue
    raise ValueError("无法读取 CSV 文件，请检查编码格式")


def _coerce_numeric_columns(df: pd.DataFrame, skip_cols: list = None) -> pd.DataFrame:
    """将数值列统一转为 float64（xlsx 可能读出 str、object、int64 或 Arrow 类型）。
    skip_cols 中的列跳过，避免误把日期列、计划 ID 列等转成数值。"""
    skip = set(skip_cols or [])
    skip.add(DATE_COL)  # 日期列永远不参与数值转换
    skip.update(ID_COLS)  # ID 列保持字符串，避免 18 位 Message ID 转 float64 丢精度
    for col in df.columns:
        if col in skip:
            continue
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].astype('float64')
        else:
            converted = pd.to_numeric(df[col], errors='coerce')
            if converted.notna().sum() > len(df) * 0.5:
                nan_count = converted.isna().sum() - df[col].isna().sum()
                if nan_count > 0:
                    logger.warning("列 '%s' 有 %d 个非数值被转为 NaN", col, nan_count)
                df[col] = converted.astype('float64')
    return df


def clean_raw_xlsx(uploaded_file) -> pd.DataFrame:
    """
    读取原始 XLSX（含 JSON 列），解析 emoji 完整保留
    1. 用 openpyxl 读取（UTF-16 内部编码，完美支持 emoji）
    2. 解析第 O 列（索引 14）的 JSON，提取标题和内容
    """
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("XLSX 文件没有数据行")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    # 列名模糊匹配（兼容 send_date 等非标准列名）
    df = _map_columns(df)
    # ID 列先定型为字符串，避免下面的数值转换把 18 位 Message ID 转成 float
    df = _normalize_id_columns(df)
    # 日期列必须先解析，再让 _coerce_numeric_columns 转换其他列
    # —— 否则 Excel 日期序列号（如 45292）会被当成 float 误转
    if DATE_COL in df.columns:
        df[DATE_COL] = _parse_date_column(df[DATE_COL])
    df = _coerce_numeric_columns(df)

    # 按列名定位 JSON 消息列（不再依赖固定位置，上游插列也不会错位）
    msg_col = _locate_message_col(df)
    if msg_col is None:
        raise ValueError("找不到 JSON 消息列，请确认文件中包含「消息内容」列")

    o_col = df[msg_col]

    # 检查公式单元格未计算的情况（data_only=True 对未缓存公式返回 None）
    if o_col.isna().all():
        raise ValueError(f"「{msg_col}」列数据全为空，可能是未计算的公式。请在 Excel 中打开文件后再上传")

    # 执行解析
    parsed_df = o_col.apply(parse_message)

    # 合并回原 DataFrame，删除原始 JSON 列
    df['标题'] = parsed_df['标题']
    df['内容'] = parsed_df['内容']
    df = df.drop(columns=[msg_col])

    return df


def read_cleaned_xlsx(uploaded_file) -> pd.DataFrame:
    """读取已清洗的 XLSX，emoji 完整保留"""
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("XLSX 文件没有数据行")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = rows[1:]

    df = pd.DataFrame(data_rows, columns=headers)
    # 列名模糊匹配 + 日期智能解析（必须先于 _coerce_numeric_columns）
    df = _map_columns(df)
    df = _normalize_id_columns(df)
    if DATE_COL in df.columns:
        df[DATE_COL] = _parse_date_column(df[DATE_COL])
    df = _coerce_numeric_columns(df)

    return df
